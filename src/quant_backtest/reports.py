from __future__ import annotations

from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .experiments import ResearchResult


CSV_OUTPUTS = {
    "base_backtest.csv": "base_backtest",
    "cost_sensitivity.csv": "cost_sensitivity",
    "parameter_sweep.csv": "parameter_sweep",
    "train_test_results.csv": "train_test_results",
    "walk_forward_results.csv": "walk_forward_results",
    "multi_asset_results.csv": "multi_asset_results",
    "model_leaderboard.csv": "model_leaderboard",
}


def save_research_outputs(result: ResearchResult, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    result.baseline_curve.to_csv(output_dir / "base_backtest.csv", index_label="Date")
    result.cost_sensitivity.to_csv(output_dir / "cost_sensitivity.csv", index=False)
    result.parameter_sweep.to_csv(output_dir / "parameter_sweep.csv", index=False)
    result.train_test_results.to_csv(output_dir / "train_test_results.csv", index=False)
    result.walk_forward_results.to_csv(output_dir / "walk_forward_results.csv", index=False)
    result.multi_asset_results.to_csv(output_dir / "multi_asset_results.csv", index=False)
    result.model_leaderboard.to_csv(output_dir / "model_leaderboard.csv", index=False)

    save_research_plots(result, output_dir)
    save_research_workbook(result, output_dir / "research_report.xlsx")


def save_research_plots(result: ResearchResult, output_dir: Path) -> None:
    _plot_baseline(result.baseline_curve, output_dir / "baseline_equity_drawdown.png")
    _plot_cost_sensitivity(result.cost_sensitivity, output_dir / "cost_sensitivity.png")
    _plot_heatmap(
        result.parameter_sweep,
        value_col="sharpe",
        title="SMA Parameter Sweep: Sharpe",
        output_path=output_dir / "sharpe_heatmap.png",
    )
    _plot_heatmap(
        result.parameter_sweep,
        value_col="cagr",
        title="SMA Parameter Sweep: CAGR",
        output_path=output_dir / "cagr_heatmap.png",
    )
    _plot_train_test(result.train_test_results, output_dir / "train_test.png")
    _plot_multi_asset(result.multi_asset_results, output_dir / "multi_asset_comparison.png")
    _plot_leaderboard(result.model_leaderboard, output_dir / "leaderboard_top_models.png")


def save_research_workbook(result: ResearchResult, output_path: Path) -> None:
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    sheets = {
        "Baseline": result.base_backtest,
        "Cost Sensitivity": result.cost_sensitivity,
        "Parameter Sweep": result.parameter_sweep,
        "Train Test": result.train_test_results,
        "Walk Forward": result.walk_forward_results,
        "Multi Asset": result.multi_asset_results,
        "Model Leaderboard": result.model_leaderboard,
        "Raw Results": _raw_results(result),
    }

    _write_dashboard(dashboard, result)
    for sheet_name, table in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        _write_table(sheet, table)

    workbook.save(output_path)


def _plot_baseline(curve: pd.DataFrame, output_path: Path) -> None:
    fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
    axes[0].plot(curve.index, curve["strategy_equity"], label="strategy")
    axes[0].plot(curve.index, curve["buy_hold_equity"], label="buy and hold")
    axes[0].set_title("Baseline equity")
    axes[0].grid(alpha=0.25)
    axes[0].legend()
    axes[1].plot(curve.index, curve["strategy_drawdown"], label="strategy drawdown")
    axes[1].plot(curve.index, curve["buy_hold_drawdown"], label="buy and hold drawdown")
    axes[1].set_title("Drawdown")
    axes[1].grid(alpha=0.25)
    axes[1].legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_cost_sensitivity(table: pd.DataFrame, output_path: Path) -> None:
    fig, axis = plt.subplots(figsize=(10, 6))
    for label, group in table.groupby("label"):
        ordered = group.sort_values("cost_bps")
        axis.plot(ordered["cost_bps"], ordered["cagr"], marker="o", label=f"{label} CAGR")
        axis.plot(ordered["cost_bps"], ordered["sharpe"], marker="s", linestyle="--", label=f"{label} Sharpe")
    axis.set_title("Transaction cost sensitivity")
    axis.set_xlabel("Cost assumption, bps")
    axis.grid(alpha=0.25)
    axis.legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_heatmap(table: pd.DataFrame, value_col: str, title: str, output_path: Path) -> None:
    pivot = table.pivot_table(index="short_window", columns="long_window", values=value_col, aggfunc="mean")
    fig, axis = plt.subplots(figsize=(8, 6))
    image = axis.imshow(pivot.values, aspect="auto", cmap="viridis")
    axis.set_xticks(range(len(pivot.columns)), labels=[str(value) for value in pivot.columns])
    axis.set_yticks(range(len(pivot.index)), labels=[str(value) for value in pivot.index])
    axis.set_xlabel("Long SMA")
    axis.set_ylabel("Short SMA")
    axis.set_title(title)
    fig.colorbar(image, ax=axis)
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_train_test(table: pd.DataFrame, output_path: Path) -> None:
    fig, axis = plt.subplots(figsize=(8, 5))
    table.plot(kind="bar", x="label", y=["cagr", "sharpe"], ax=axis)
    axis.set_title("Train vs test performance")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_multi_asset(table: pd.DataFrame, output_path: Path) -> None:
    fig, axis = plt.subplots(figsize=(11, 6))
    table.plot(kind="bar", x="ticker", y=["cagr", "benchmark_cagr"], ax=axis)
    axis.set_title("Multi-asset CAGR comparison")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_leaderboard(table: pd.DataFrame, output_path: Path) -> None:
    fig, axis = plt.subplots(figsize=(11, 6))
    top = table.head(10)
    top.plot(kind="bar", x="variant", y=["sharpe", "cagr"], ax=axis)
    axis.set_title("Top model variants")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _write_dashboard(sheet, result: ResearchResult) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "AAPL SMA Research Framework v2"
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0F172A")
    sheet.merge_cells("A1:H1")

    baseline = result.base_backtest.iloc[0]
    best = result.model_leaderboard.iloc[0]
    summary = [
        ["Baseline CAGR", baseline["cagr"]],
        ["Baseline Sharpe", baseline["sharpe"]],
        ["Baseline Max Drawdown", baseline["max_drawdown"]],
        ["Best Variant", best["variant"]],
        ["Best Variant CAGR", best["cagr"]],
        ["Best Variant Sharpe", best["sharpe"]],
    ]
    for row_idx, row in enumerate(summary, start=3):
        sheet.cell(row_idx, 1, row[0])
        sheet.cell(row_idx, 2, row[1])
    _write_table(sheet, result.model_leaderboard.head(12), start_row=11, table_name="DashboardLeaderboard")
    _add_dashboard_chart(sheet, result)
    _set_widths(sheet, [24, 18, 16, 16, 16, 16, 16, 16])


def _add_dashboard_chart(sheet, result: ResearchResult) -> None:
    start_row = 26
    table = result.cost_sensitivity[["label", "cost_bps", "cagr"]].copy()
    table["scenario"] = table["label"] + " " + table["cost_bps"].astype(str) + "bps"
    compact = table[["scenario", "cagr"]]
    _write_table(sheet, compact, start_row=start_row, table_name="DashboardCosts")
    chart = BarChart()
    chart.title = "Cost sensitivity CAGR"
    data = Reference(sheet, min_col=2, min_row=start_row, max_row=start_row + len(compact))
    cats = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(compact))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    sheet.add_chart(chart, "D26")


def _write_table(sheet, table: pd.DataFrame, start_row: int = 1, table_name: str | None = None) -> None:
    if table.empty:
        sheet.cell(start_row, 1, "No rows")
        return
    values = [list(table.columns)] + table.replace({pd.NA: None}).where(pd.notna(table), None).values.tolist()
    for row_offset, row in enumerate(values):
        for col_offset, value in enumerate(row):
            cell = sheet.cell(start_row + row_offset, 1 + col_offset, value)
            if row_offset == 0:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill("solid", fgColor="111827")
    table_ref = (
        f"A{start_row}:"
        f"{get_column_letter(len(table.columns))}{start_row + len(table)}"
    )
    name = table_name or sheet.title.replace(" ", "") + "Table"
    excel_table = Table(displayName=_safe_table_name(name), ref=table_ref)
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(excel_table)
    _set_widths(sheet, [max(12, min(28, len(str(col)) + 4)) for col in table.columns])


def _raw_results(result: ResearchResult) -> pd.DataFrame:
    return pd.concat(
        [
            result.cost_sensitivity.assign(source="cost_sensitivity"),
            result.parameter_sweep.assign(source="parameter_sweep"),
            result.train_test_results.assign(source="train_test"),
            result.walk_forward_results.assign(source="walk_forward"),
            result.multi_asset_results.assign(source="multi_asset"),
            result.model_leaderboard.assign(source="leaderboard"),
        ],
        ignore_index=True,
        sort=False,
    )


def _set_widths(sheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _safe_table_name(name: str) -> str:
    return "".join(char for char in name if char.isalnum() or char == "_")[:240]
