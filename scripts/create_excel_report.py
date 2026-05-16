from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import AreaChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs"
MODEL_DATA_CSV = OUTPUT_DIR / "excel_model_data.csv"
METRICS_CSV = OUTPUT_DIR / "metrics.csv"
WORKBOOK_OUT = OUTPUT_DIR / "aapl_model_forecast_report.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="111827")
TITLE_FILL = PatternFill("solid", fgColor="0F172A")
LABEL_FILL = PatternFill("solid", fgColor="E5E7EB")
VALUE_FILL = PatternFill("solid", fgColor="F8FAFC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True)


def main() -> None:
    curve = pd.read_csv(MODEL_DATA_CSV, parse_dates=["Date"])
    metrics = pd.read_csv(METRICS_CSV)

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    data_sheet = workbook.create_sheet("Model Data")
    metrics_sheet = workbook.create_sheet("Metrics")

    write_data_sheet(data_sheet, curve)
    write_metrics_sheet(metrics_sheet, metrics)
    write_dashboard(dashboard, curve, metrics)

    workbook.save(WORKBOOK_OUT)
    print(f"Wrote workbook: {WORKBOOK_OUT}")


def write_data_sheet(sheet, curve: pd.DataFrame) -> None:
    columns = [
        ("Date", "Date"),
        ("Price", "price"),
        ("SMA 20", "short_sma"),
        ("SMA 100", "long_sma"),
        ("Model Signal", "signal"),
        ("Executed Position", "position"),
        ("Strategy Equity", "strategy_equity"),
        ("Buy Hold Equity", "buy_hold_equity"),
        ("Entry Price", "entry_price"),
        ("Exit Price", "exit_price"),
        ("Model Forecast", "model_forecast"),
    ]
    sheet.append([header for header, _ in columns])

    for _, row in curve.iterrows():
        sheet.append([clean_excel_value(row[source]) for _, source in columns])

    style_header(sheet, 1, len(columns))
    set_column_widths(sheet, [13, 12, 12, 12, 13, 16, 16, 16, 12, 12, 15])
    add_table(sheet, "ModelDataTable", 1, 1, sheet.max_row, len(columns))
    sheet.freeze_panes = "A2"

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        for col in [2, 3, 4, 9, 10]:
            sheet.cell(row, col).number_format = "$0.00"
        for col in [7, 8]:
            sheet.cell(row, col).number_format = "$#,##0"


def write_metrics_sheet(sheet, metrics: pd.DataFrame) -> None:
    headers = ["Model", "Total Return", "CAGR", "Volatility", "Sharpe", "Max Drawdown", "Trades", "Win Rate"]
    sheet.append(headers)
    for _, row in metrics.iterrows():
        sheet.append(
            [
                row["name"],
                row["total_return"],
                row["cagr"],
                row["ann_volatility"],
                row["sharpe"],
                row["max_drawdown"],
                row["trades"],
                clean_excel_value(row["win_rate"]),
            ]
        )

    style_header(sheet, 1, len(headers))
    set_column_widths(sheet, [14, 14, 12, 12, 10, 14, 10, 10])
    add_table(sheet, "MetricsTable", 1, 1, sheet.max_row, len(headers))

    for row in range(2, sheet.max_row + 1):
        for col in [2, 3, 4, 6, 8]:
            sheet.cell(row, col).number_format = "0.00%"
        sheet.cell(row, 5).number_format = "0.00"


def write_dashboard(sheet, curve: pd.DataFrame, metrics: pd.DataFrame) -> None:
    metric_map = metrics.set_index("name")
    latest = curve.iloc[-1]

    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "AAPL SMA Model Forecast Report"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    sheet["A1"].alignment = Alignment(vertical="center")

    sheet.merge_cells("A2:H2")
    sheet["A2"] = "Historical price, SMA regime forecasts, executed trades, and strategy vs buy-and-hold performance."
    sheet["A2"].font = Font(color="334155", italic=True)

    kpis = [
        ("Latest date", latest["Date"].date()),
        ("Latest price", latest["price"]),
        ("Current model forecast", latest["model_forecast"]),
        ("Strategy CAGR", metric_map.loc["strategy", "cagr"]),
        ("Buy & hold CAGR", metric_map.loc["buy_hold", "cagr"]),
        ("Strategy Sharpe", metric_map.loc["strategy", "sharpe"]),
    ]
    for idx, (label, value) in enumerate(kpis, start=4):
        sheet.cell(idx, 1, label)
        sheet.cell(idx, 2, clean_excel_value(value))
        sheet.cell(idx, 1).fill = LABEL_FILL
        sheet.cell(idx, 1).font = Font(bold=True)
        sheet.cell(idx, 2).fill = VALUE_FILL

    sheet["B4"].number_format = "yyyy-mm-dd"
    sheet["B5"].number_format = "$0.00"
    sheet["B7"].number_format = "0.00%"
    sheet["B8"].number_format = "0.00%"
    sheet["B9"].number_format = "0.00"

    comparison = [
        ["Metric", "Strategy", "Buy & Hold", "Difference", "Interpretation"],
        [
            "Total return",
            metric_map.loc["strategy", "total_return"],
            metric_map.loc["buy_hold", "total_return"],
            metric_map.loc["strategy", "total_return"] - metric_map.loc["buy_hold", "total_return"],
            "Higher is better",
        ],
        [
            "Max drawdown",
            metric_map.loc["strategy", "max_drawdown"],
            metric_map.loc["buy_hold", "max_drawdown"],
            metric_map.loc["strategy", "max_drawdown"] - metric_map.loc["buy_hold", "max_drawdown"],
            "Less negative is better",
        ],
    ]
    for row_idx, row_values in enumerate(comparison, start=4):
        for col_idx, value in enumerate(row_values, start=4):
            sheet.cell(row_idx, col_idx, clean_excel_value(value))
    style_header(sheet, 4, 8, start_col=4)
    for row in [5, 6]:
        for col in [5, 6, 7]:
            sheet.cell(row, col).number_format = "0.00%"

    helper_start = 10
    helper_headers = ["Date", "Price", "SMA 20", "SMA 100", "Signal Date", "Model Signal"]
    for col_idx, header in enumerate(helper_headers, start=helper_start):
        cell = sheet.cell(1, col_idx, header)
        cell.font = Font(color="FFFFFF")
    for row_idx, (_, row) in enumerate(curve.iterrows(), start=2):
        sheet.cell(row_idx, helper_start, row["Date"])
        sheet.cell(row_idx, helper_start + 1, clean_excel_value(row["price"]))
        sheet.cell(row_idx, helper_start + 2, clean_excel_value(row["short_sma"]))
        sheet.cell(row_idx, helper_start + 3, clean_excel_value(row["long_sma"]))
        sheet.cell(row_idx, helper_start + 4, row["Date"])
        sheet.cell(row_idx, helper_start + 5, clean_excel_value(row["signal"]))

    max_row = len(curve) + 1
    add_price_chart(sheet, helper_start, max_row)
    add_forecast_chart(sheet, helper_start + 4, max_row)
    set_column_widths(sheet, [18, 15, 3, 16, 15, 15, 14, 18])
    for col in range(helper_start, helper_start + len(helper_headers)):
        sheet.column_dimensions[get_column_letter(col)].hidden = True


def add_price_chart(sheet, helper_start: int, max_row: int) -> None:
    chart = LineChart()
    chart.title = "AAPL price with SMA model inputs"
    chart.height = 9
    chart.width = 24
    chart.y_axis.title = "Price"
    chart.x_axis.title = "Date"
    chart.x_axis.tickLblSkip = 252
    chart.y_axis.majorGridlines = ChartLines()

    data = Reference(sheet, min_col=helper_start + 1, max_col=helper_start + 3, min_row=1, max_row=max_row)
    categories = Reference(sheet, min_col=helper_start, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "A11")


def add_forecast_chart(sheet, helper_start: int, max_row: int) -> None:
    chart = AreaChart()
    chart.title = "Model forecast regime: 1 = LONG, 0 = CASH"
    chart.height = 7
    chart.width = 24
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.5
    chart.x_axis.tickLblSkip = 252

    data = Reference(sheet, min_col=helper_start + 1, min_row=1, max_row=max_row)
    categories = Reference(sheet, min_col=helper_start, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "A31")


def style_header(sheet, row: int, end_col: int, start_col: int = 1) -> None:
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def set_column_widths(sheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def add_table(sheet, name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)


def clean_excel_value(value):
    if pd.isna(value):
        return None
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    return value


if __name__ == "__main__":
    main()
